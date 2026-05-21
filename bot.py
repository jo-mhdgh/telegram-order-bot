from keep_alive import keep_alive
import os
import math
import sqlite3
from datetime import datetime, date, timezone, timedelta
from dotenv import load_dotenv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

LOCAL_TZ = timezone(timedelta(hours=3))

from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.ext import (
    ApplicationBuilder,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

# =========================
# LOAD ENV
# =========================

load_dotenv()
BOT_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID"))

# =========================
# CONFIG
# =========================

OPEN_HOUR = 8
CLOSE_HOUR = 17

PRODUCTS = [
    ("فطور", "🍳"),
    ("غداء", "🍽"),
    ("عشاء", "🥫"),
    ("شاي", "☕️"),
    ("ماء", "💧"),
]

# =========================
# DATABASE
# =========================

conn = sqlite3.connect("orders.db", check_same_thread=False)
cursor = conn.cursor()

cursor.execute("""
CREATE TABLE IF NOT EXISTS orders (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER,
    username TEXT,
    order_name TEXT,
    product TEXT,
    quantity INTEGER,
    datetime TEXT
)
""")

conn.commit()




# =========================
# MEMORY
# =========================

user_carts = {}
waiting_for_name = {}

# =========================
# HELPERS
# =========================

def is_open():
    now = datetime.now(LOCAL_TZ).hour
    return OPEN_HOUR <= now < CLOSE_HOUR

def today():
    return datetime.now(LOCAL_TZ).date().isoformat()

# =========================
# KEYBOARD
# =========================

def build_keyboard(user_id):
    cart = user_carts.get(user_id, {})
    text = "🛒 طلبك:\n\n"
    keyboard = []

    for product, emoji in PRODUCTS:
        qty = cart.get(product, 0)
        text += f"{emoji} {product}: {qty}\n"


        keyboard.append([
            InlineKeyboardButton(f"{emoji} {product}", callback_data="noop"),
        ])

        keyboard.append([
            InlineKeyboardButton("-1", callback_data=f"minus:{product}"),
            InlineKeyboardButton("+1", callback_data=f"plus1:{product}"),
            InlineKeyboardButton("+5", callback_data=f"plus5:{product}"),
        ])
    keyboard.append([
        InlineKeyboardButton("✅ تأكيد الطلب", callback_data="confirm")
    ])

    keyboard.append([
        InlineKeyboardButton("🗑 حذف", callback_data="clear")
    ])

    return text, InlineKeyboardMarkup(keyboard)

# =========================
# START
# =========================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):

    user_id = update.effective_user.id

    if not is_open():
        await update.message.reply_text("⛔ الطلبات مغلقة (8ص - 5م)")
        return

    # ✅ CHECK IF ALREADY ORDERED
    cursor.execute("""
    SELECT product, quantity
    FROM orders
    WHERE user_id=? AND DATE(datetime)=?
    """, (user_id, today()))

    rows = cursor.fetchall()

    if rows:
        text = "📦 طلبك اليوم:\n\n"

        for product, qty in rows:
            text += f"{product}: {qty}\n"

        await update.message.reply_text(text)
        return

    # ✅ normal flow
    user_carts[user_id] = {}

    text, kb = build_keyboard(user_id)
    await update.message.reply_text(text, reply_markup=kb)
    
# =========================
# BUTTONS
# =========================

async def button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id

    if not is_open():
        await query.answer("⛔ مغلق الآن", show_alert=True)
        return

    if user_id not in user_carts:
        user_carts[user_id] = {}

    cart = user_carts[user_id]
    data = query.data

    if data.startswith("plus1:"):
        p = data.split(":")[1]
        cart[p] = cart.get(p, 0) + 1

    elif data.startswith("plus5:"):
        p = data.split(":")[1]
        cart[p] = cart.get(p, 0) + 5

    elif data.startswith("minus:"):
        p = data.split(":")[1]
        if cart.get(p, 0) > 0:
            cart[p] -= 1

    elif data == "clear":
        user_carts[user_id] = {}

    elif data == "confirm":

        if not any(cart.values()):
            await query.answer("السلة فارغة", show_alert=True)
            return

        cursor.execute("""
        SELECT COUNT(*) FROM orders
        WHERE user_id=? AND DATE(datetime)=?
        """, (user_id, today()))

        if cursor.fetchone()[0] > 0:
            await query.edit_message_text("❌ لقد قمت بإرسال طلبك اليوم بالفعل")
            return

        waiting_for_name[user_id] = True
        await query.edit_message_text("📝 اكتب اسم الطلب (إجباري):")
        return

    text, kb = build_keyboard(user_id)
    await query.edit_message_text(text, reply_markup=kb)

# =========================
# RECEIVE NAME
# =========================
async def receive_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    if user_id not in waiting_for_name:
        return

    order_name = update.message.text
    cart = user_carts.get(user_id, {})

    username = update.effective_user.username or "NoUsername"

    # ✅ SAVE TO DATABASE
    for product, qty in cart.items():
        if qty > 0:
            cursor.execute("""
            INSERT INTO orders (user_id, username, order_name, product, quantity, datetime)
            VALUES (?, ?, ?, ?, ?, ?)
            """, (
                user_id,
                username,
                order_name,
                product,
                qty,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ))

    conn.commit()

    # =========================
    # ✅ SEND TO ADMIN (INSIDE FUNCTION)
    # =========================
    admin_text = f"📥 طلب جديد:\n\n"
    admin_text += f"👤 @{username}\n"
    admin_text += f"🧾 {order_name}\n"
    admin_text += f"⏰ {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"

    for p, q in cart.items():
        if q > 0:
            admin_text += f"{p}: {q}\n"

    try:
        await context.bot.send_message(chat_id=ADMIN_ID, text=admin_text)
    except Exception as e:
        print(f"Failed to notify admin: {e}")

    # =========================
    # ✅ USER CONFIRMATION
    # =========================
    summary = f"✅ تم تسجيل: {order_name}\n\n"
    for p, q in cart.items():
        if q > 0:
            summary += f"{p}: {q}\n"

    user_carts[user_id] = {}
    waiting_for_name.pop(user_id)

    await update.message.reply_text(summary)

# =========================
# SUMMARY
# =========================

async def summary(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return

    cursor.execute("""
    SELECT product, SUM(quantity)
    FROM orders
    WHERE DATE(datetime)=?
    GROUP BY product
    """, (today(),))

    rows = cursor.fetchall()

    text = "📊 ملخص اليوم:\n\n"
    for p, total in rows:
        text += f"{p}: {total}\n"

    await update.message.reply_text(text)
# =========================
# DETAILS (NEW)
# =========================
async def details(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return

    cursor.execute("""
    SELECT order_name, username, product, quantity, datetime
    FROM orders
    WHERE DATE(datetime)=?
    ORDER BY datetime
    """, (today(),))

    rows = cursor.fetchall()

    if not rows:
        await update.message.reply_text("لا توجد طلبات اليوم")
        return

    orders = {}

    for order_name, username, product, quantity, dt in rows:
        key = f"{order_name} (@{username}) | {dt}"

        if key not in orders:
            orders[key] = []

        orders[key].append((product, quantity))

    text = "📋 تفاصيل الطلبات:\n\n"

    for order_name, items in orders.items():
        text += f"🧾 {order_name}\n"
        for product, qty in items:
            text += f"  - {product}: {qty}\n"
        text += "\n"

    if len(text) > 4000:
        await update.message.reply_text("البيانات كبيرة، استخدم /export")
        return

    await update.message.reply_text(text)

# =========================
# EXCEL REPORT BUILDER
# =========================

def round_up_even(n):
    return math.ceil(n / 2) * 2

def generate_excel_report(rows, file_name, title="تقرير الطلبات"):
    wb = Workbook()

    # ── Styles ──────────────────────────────────────────────
    hdr_fill   = PatternFill("solid", fgColor="1F4E79")
    grp_fill   = PatternFill("solid", fgColor="2E75B6")
    sub_fill   = PatternFill("solid", fgColor="D6E4F0")
    sum_fill   = PatternFill("solid", fgColor="E2EFDA")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    alt_fill   = PatternFill("solid", fgColor="F2F7FB")

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    grp_font  = Font(bold=True, color="FFFFFF", size=11)
    sub_font  = Font(bold=True, color="1F4E79", size=10)
    body_font = Font(size=10)
    title_font= Font(bold=True, size=14, color="1F4E79")

    thin = Side(style="thin", color="BDD7EE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    # ── Sheet 1: Detailed Orders ─────────────────────────────
    ws = wb.active
    ws.title = "الطلبات التفصيلية"
    ws.sheet_view.rightToLeft = True

    # Title row
    ws.merge_cells("A1:E1")
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    # Header row
    headers = ["الاسم", "المستخدم", "المنتج", "الكمية (معدّلة)", "الوقت"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 20

    # Group rows by order_name
    from collections import defaultdict
    groups = defaultdict(list)
    for order_name, username, product, quantity, dt in rows:
        groups[order_name].append((username, product, quantity, dt))

    current_row = 3
    product_totals = defaultdict(int)

    for idx, (order_name, items) in enumerate(groups.items()):
        # Group header
        ws.merge_cells(f"A{current_row}:E{current_row}")
        cell = ws.cell(row=current_row, column=1, value=f"  {order_name}")
        cell.font = grp_font
        cell.fill = grp_fill
        cell.alignment = left
        cell.border = border
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        group_totals = defaultdict(int)
        for i, (username, product, quantity, dt) in enumerate(items):
            adj_qty = round_up_even(quantity)
            product_totals[product] += adj_qty
            group_totals[product] += adj_qty

            fill = white_fill if i % 2 == 0 else alt_fill
            values = [order_name, f"@{username}", product, adj_qty, dt]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.font = body_font
                cell.fill = fill
                cell.alignment = center if col != 1 else left
                cell.border = border
            current_row += 1

        # Subtotal row per group
        ws.merge_cells(f"A{current_row}:B{current_row}")
        sub_label = ws.cell(row=current_row, column=1, value="المجموع الفرعي")
        sub_label.font = sub_font
        sub_label.fill = sub_fill
        sub_label.alignment = center
        sub_label.border = border
        ws.cell(row=current_row, column=2).fill = sub_fill
        ws.cell(row=current_row, column=2).border = border

        products_str = "  |  ".join(f"{p}: {q}" for p, q in group_totals.items())
        total_qty = sum(group_totals.values())

        prod_cell = ws.cell(row=current_row, column=3, value=products_str)
        prod_cell.font = sub_font
        prod_cell.fill = sub_fill
        prod_cell.alignment = left
        prod_cell.border = border

        qty_cell = ws.cell(row=current_row, column=4, value=total_qty)
        qty_cell.font = sub_font
        qty_cell.fill = sub_fill
        qty_cell.alignment = center
        qty_cell.border = border

        ws.cell(row=current_row, column=5).fill = sub_fill
        ws.cell(row=current_row, column=5).border = border
        ws.row_dimensions[current_row].height = 18
        current_row += 2  # blank gap between groups

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20

    # ── Sheet 2: Grand Summary ───────────────────────────────
    ws2 = wb.create_sheet("ملخص الكميات")
    ws2.sheet_view.rightToLeft = True

    ws2.merge_cells("A1:B1")
    ws2["A1"] = "ملخص إجمالي للكميات"
    ws2["A1"].font = title_font
    ws2["A1"].alignment = center
    ws2.row_dimensions[1].height = 28

    for col, h in enumerate(["المنتج", "الكمية الإجمالية"], 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
    ws2.row_dimensions[2].height = 20

    grand_total = 0
    for i, (product, total) in enumerate(product_totals.items()):
        fill = white_fill if i % 2 == 0 else alt_fill
        ws2.cell(row=3+i, column=1, value=product).fill = fill
        ws2.cell(row=3+i, column=1).font = body_font
        ws2.cell(row=3+i, column=1).alignment = center
        ws2.cell(row=3+i, column=1).border = border
        ws2.cell(row=3+i, column=2, value=total).fill = fill
        ws2.cell(row=3+i, column=2).font = body_font
        ws2.cell(row=3+i, column=2).alignment = center
        ws2.cell(row=3+i, column=2).border = border
        grand_total += total

    total_row = 3 + len(product_totals)
    ws2.cell(row=total_row, column=1, value="الإجمالي الكلي").font = Font(bold=True, size=11, color="1F4E79")
    ws2.cell(row=total_row, column=1).fill = sum_fill
    ws2.cell(row=total_row, column=1).alignment = center
    ws2.cell(row=total_row, column=1).border = border
    ws2.cell(row=total_row, column=2, value=grand_total).font = Font(bold=True, size=11, color="1F4E79")
    ws2.cell(row=total_row, column=2).fill = sum_fill
    ws2.cell(row=total_row, column=2).alignment = center
    ws2.cell(row=total_row, column=2).border = border
    ws2.row_dimensions[total_row].height = 20

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 22

    wb.save(file_name)

# =========================
# EXPORT
# =========================

async def export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return

    if context.args:
        target_date = context.args[0]
        try:
            datetime.strptime(target_date, "%Y-%m-%d")
        except ValueError:
            await update.message.reply_text("❌ صيغة التاريخ خاطئة. استخدم: /export YYYY-MM-DD")
            return
        label = target_date
    else:
        target_date = today()
        label = "اليوم"

    cursor.execute("""
    SELECT order_name, username, product, quantity, datetime
    FROM orders
    WHERE DATE(datetime)=?
    """, (target_date,))

    rows = cursor.fetchall()

    if not rows:
        await update.message.reply_text(f"لا يوجد بيانات في {label}")
        return

    file_name = f"orders_{target_date}.xlsx"
    generate_excel_report(rows, file_name, title=f"تقرير الطلبات — {target_date}")

    await update.message.reply_document(open(file_name, "rb"))

# =========================
# EXPORT ALL
# =========================

async def export_all(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return

    cursor.execute("""
    SELECT order_name, username, product, quantity, datetime
    FROM orders
    ORDER BY datetime
    """)

    rows = cursor.fetchall()

    if not rows:
        await update.message.reply_text("لا يوجد بيانات")
        return

    file_name = "orders_all.xlsx"
    generate_excel_report(rows, file_name, title="تقرير جميع الطلبات")

    await update.message.reply_document(open(file_name, "rb"))

# =========================
# RESET
# =========================

async def reset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return

    cursor.execute("DELETE FROM orders")
    conn.commit()

    await update.message.reply_text("🗑 تم تصفير اليوم")

# =========================
# MAIN
# =========================

def main():
    app = ApplicationBuilder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("summary", summary))
    app.add_handler(CommandHandler("details", details))  # 👈 NEW
    app.add_handler(CommandHandler("export", export_excel))
    app.add_handler(CommandHandler("exportall", export_all))
    app.add_handler(CommandHandler("reset", reset))

    app.add_handler(CallbackQueryHandler(button))
    app.add_handler(MessageHandler(filters.TEXT, receive_name))

    print("Bot running...")
    keep_alive()
    app.run_polling()

if __name__ == "__main__":
    main()
